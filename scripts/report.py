from openpyxl.utils import get_column_letter
import copy

def copy_style(from_cell, to_cell):
    if from_cell.has_style:
        to_cell.font = copy.copy(from_cell.font)
        to_cell.border = copy.copy(from_cell.border)
        to_cell.fill = copy.copy(from_cell.fill)
        to_cell.number_format = from_cell.number_format
        to_cell.protection = copy.copy(from_cell.protection)
        to_cell.alignment = copy.copy(from_cell.alignment)

def update_excel_table(wb, sheet_name, table_name, data, start_col=3):
    ws = wb[sheet_name]

    # 取得表格起始資訊
    table = ws.tables[table_name]
    table.ref = table.ref or table.tableRef
    start_cell = table.ref.split(':')[0]  # e.g. C9
    col_letter = ''.join(filter(str.isalpha, start_cell))  # 'C'
    row_number = int(''.join(filter(str.isdigit, start_cell)))  # 9

    # 資料區開始列與資料尺寸
    data_start_row = row_number + 1
    num_cols = len(data[0])
    num_rows = len(data)

    # ✅ Step 3: 取得樣式來源格（原 C10）
    style_source_cell = ws.cell(row=data_start_row, column=start_col)
    style_left_cell = ws.cell(row=data_start_row, column=start_col-1)
    style_right_cell = ws.cell(row=data_start_row, column=start_col + len(data[0]))# ⚠ 來源往後抓，以免剛刪掉
    
    # ✅ Step 1: 刪掉假資料的前 2 行
    # 清除原資料列（這裡假設原資料不超過 50 列）
    for _ in range(2):
        ws.delete_rows(data_start_row)


    # ✅ Step 2: 插入空白列，讓底線與格式自動下移
    ws.insert_rows(data_start_row, amount=num_rows)


    # ✅ Step 4: 寫入資料並套用樣式
    for i, row_data in enumerate(data):
        cell_left = ws.cell(row=data_start_row + i, column=start_col -1, value="")
        copy_style(style_left_cell, cell_left)
        cell_right = ws.cell(row=data_start_row + i, column=start_col + len(row_data), value="")
        copy_style(style_right_cell, cell_right)
        for j, val in enumerate(row_data):
            cell = ws.cell(row=data_start_row + i, column=start_col + j, value=val)
            copy_style(style_source_cell, cell)

    # ✅ Step 5: 更新 table.ref 範圍
    end_col_letter = get_column_letter(start_col + num_cols - 1)
    end_row = data_start_row + num_rows - 1
    new_ref = f"{col_letter}{row_number}:{end_col_letter}{end_row}"
    table.ref = new_ref

    # 儲存結果
    return wb
    #wb.save(output_path)
    #print(f"✅ 表格 {table_name} 已更新並輸出至 {output_path}")
    
    
def update_excel_table_with_chart(wb, sheet_name, table_name, data, metadata, start_col=3):
    ws = wb[sheet_name]
    update_metadata_cells(wb, sheet_name, metadata)

    # 取得表格起始資訊
    table = ws.tables[table_name]
    table.ref = table.ref or table.tableRef
    start_cell = table.ref.split(':')[0]
    col_letter = ''.join(filter(str.isalpha, start_cell))
    row_number = int(''.join(filter(str.isdigit, start_cell)))

    # 資料區開始列與資料尺寸
    data_start_row = row_number + 1
    num_cols = len(data[0])
    num_rows = len(data)

    # 樣式來源儲存格
    style_source_cell = ws.cell(row=data_start_row, column=start_col)
    style_left_cell = ws.cell(row=data_start_row, column=start_col-1)
    style_right_cell = ws.cell(row=data_start_row, column=start_col + 8)

    # 刪除假資料列
    for _ in range(2):
        ws.delete_rows(data_start_row)

    # 插入新資料所需空白列
    ws.insert_rows(data_start_row, amount=num_rows)

    # 寫入資料與套用樣式
    for i, row_data in enumerate(data):
        cell_left = ws.cell(row=data_start_row + i, column=start_col -1, value="")
        copy_style(style_left_cell, cell_left)
        cell_right = ws.cell(row=data_start_row + i, column=start_col + 8, value="")
        copy_style(style_right_cell, cell_right)
        for j, val in enumerate(row_data):
            cell = ws.cell(row=data_start_row + i, column=start_col + j, value=val)
            cell2 = ws.cell(row=data_start_row + i, column=start_col + j + 4, value="")
            copy_style(style_source_cell, cell)
            copy_style(style_source_cell, cell2)

    # 更新表格範圍
    end_col_letter = get_column_letter(start_col + num_cols - 1)
    end_row = data_start_row + num_rows - 1
    new_ref = f"{col_letter}{row_number}:{end_col_letter}{end_row}"
    table.ref = new_ref

    # 更新圖表資料範圍
    charts = ws._charts
    for idx, chart in enumerate(charts):
        if idx < num_cols - 1:  # 假設每個圖表對應一欄資料
            value_col = get_column_letter(start_col + 2 + idx)
            category_col = get_column_letter(start_col)
            chart.series[0].val.numRef.f = f"'{sheet_name}'!${value_col}${data_start_row}:${value_col}${end_row}"
            chart.series[0].cat.strRef.f = f"'{sheet_name}'!${category_col}${data_start_row}:${category_col}${end_row}"

    # 儲存結果
    return wb
    #wb.save(output_path)
    #print(f"✅ 表格 {table_name} 與圖表已更新並輸出至 {output_path}")
    
def update_metadata_cells(wb, sheet_name, metadata):
    """
    wb: openpyxl Workbook 物件
    sheet_name: 工作表名稱
    metadata: dict，格式為 {'Department': ..., 'Manager': ..., 'Start Time': ..., 'End Time': ...}
    """
    print("有進來喔")

    ws = wb[sheet_name]
    keys = ['Team', 'Manager', 'Start Time', 'End Time']
    start_row = 9
    col = 4  # D 欄

    for i, key in enumerate(keys):
        cell = ws.cell(row=start_row + i, column=col)
        if key in metadata:
            cell.value = metadata[key]




    
#data = [
#    ['SOP001', '建立流程', '2025-04-26 16:44:18', 'v1.0'],
#    ['SOP002', '審核流程', '2025-04-26 16:44:18', 'v1.1'],
#    ['SOP002', '審核流程', '2025-04-26 16:44:18', 'v1.1'],
#    ['SOP002', '審核流程', '2025-04-26 16:44:18', 'v1.1'],
#    ['SOP002', '審核流程', '2025-04-26 16:44:18', 'v1.1']
#]

#data = [['sop001', 'yen', 10, 20],['sop002', 'yen', 10, 20],['sop003', 'yen', 10, 20],['sop004', 'yen', 10, 20],['sop005', 'yen', 10, 20],['sop006', 'yen', 10, 20]]
#metadata = {
#    "Department": "IT Department",
#    "Manager": "Yen Wu",
#    "Start Time": "2025-05-21",
#    "End Time": "2025-05-21"
#}
#update_excel_table_with_chart(
#    file_path='/Users/yen/NCCU/大四（二）/DBMS/113-2-DBMS-Final-Project/src/assets/ReportTemplate.xlsx',
#    output_path='output.xlsx',
#    sheet_name='Social Interaction',
#    table_name='Table3',
#    data=data,
#    metadata = metadata,
#    start_col=3  # C 欄
#)
